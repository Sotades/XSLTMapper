"""Module for reading Excel files using openpyxl."""

from pathlib import Path
from typing import Union, List, Dict, Any, Optional
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


class ExcelReader:
    """A class to handle reading Excel files with proper resource management."""
    
    def __init__(self, file_path: Union[str, Path]) -> None:
        """Initialize the Excel reader with a file path.
        
        Args:
            file_path: Path to the Excel file (string or Path object)
        
        Raises:
            FileNotFoundError: If the specified file doesn't exist
        """
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.file_path}")
        
        self.workbook: Optional[Workbook] = None
    
    def __enter__(self) -> 'ExcelReader':
        """Context manager entry method.
        
        Returns:
            ExcelReader instance
        """
        self.workbook = openpyxl.load_workbook(
            filename=self.file_path,
            read_only=True,
            data_only=True
        )
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb) -> None:
        """Context manager exit method for cleanup."""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
    
    def read_worksheet(
        self,
        worksheet_name: str,
        header: bool = True,
        start_row: int = 1
    ) -> List[Dict[str, Any]]:
        """Read a worksheet and return its data as a list of dictionaries.
        
        Args:
            worksheet_name: Name of the worksheet to read
            header: Whether to use the first row as headers (default: True)
            start_row: Row number to start reading from (default: 1)
        
        Returns:
            List of dictionaries where keys are column headers and values are cell values
        
        Raises:
            ValueError: If worksheet name doesn't exist
            RuntimeError: If workbook is not open
        """
        if not self.workbook:
            self.workbook = openpyxl.load_workbook(
                filename=self.file_path,
                read_only=True,
                data_only=True
            )
        
        if worksheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Worksheet '{worksheet_name}' not found in workbook")
        
        worksheet = self.workbook[worksheet_name]
        rows = list(worksheet.rows)
        
        if not rows:
            return []
        
        # Handle headers
        if header:
            headers = [
                str(cell.value) if cell.value is not None else f"Column{i+1}"
                for i, cell in enumerate(rows[start_row - 1])
            ]
            start_row += 1
        else:
            headers = [
                f"Column{i+1}" 
                for i in range(1, worksheet.max_column + 1)
            ]
        
        # Read data rows
        data = []
        for row in rows[start_row - 1:]:
            row_data = {
                headers[i]: cell.value
                for i, cell in enumerate(row)
                if i < len(headers)
            }
            data.append(row_data)
        
        return data
    
    def get_worksheet_names(self) -> List[str]:
        """Get the names of all worksheets in the workbook.
        
        Returns:
            List of worksheet names
        """
        if not self.workbook:
            self.workbook = openpyxl.load_workbook(
                filename=self.file_path,
                read_only=True,
                data_only=True
            )
        
        return self.workbook.sheetnames
    
    def close(self) -> None:
        """Close the workbook and release resources."""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
